"""
Report Generation Module - Excel and PDF Export
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from typing import Dict, Any, List
from datetime import datetime
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generate Excel and PDF reports"""

    @staticmethod
    def generate_excel_report(project, calculations: List[Dict], output_path: Path) -> bool:
        """
        Generate Excel report for a project

        Args:
            project: Project object
            calculations: List of calculation dictionaries
            output_path: Output file path

        Returns:
            bool: True if successful
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "GHG Emissions Report"

            # Header
            ws['A1'] = "GHG EMISSIONS REPORT"
            ws['A1'].font = Font(size=16, bold=True)

            # Project info
            row = 3
            info_data = [
                ["Project ID:", project.id],
                ["Project Name:", project.project_name],
                ["Organization:", project.organization_name],
                ["Reporting Year:", project.reporting_year],
                ["Report Date:", datetime.now().strftime("%Y-%m-%d")],
                ["Status:", project.status]
            ]

            for label, value in info_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

            # Emissions summary
            row += 2
            ws[f'A{row}'] = "EMISSIONS SUMMARY"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 2

            # Summary table headers
            headers = ["Scope", "Category", "Activity Data", "Emission Factor", "Emissions (tCO2e)"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            row += 1

            # Add calculations
            total_emissions = 0
            for calc in calculations:
                ws.cell(row=row, column=1, value=calc.get('scope', 'N/A'))
                ws.cell(row=row, column=2, value=calc.get('category', 'N/A'))
                ws.cell(row=row, column=3, value=calc.get('activity_data', 0))
                ws.cell(row=row, column=4, value=calc.get('emission_factor', 0))
                ws.cell(row=row, column=5, value=calc.get('emissions_tco2e', 0))
                total_emissions += calc.get('emissions_tco2e', 0)
                row += 1

            # Total row
            row += 1
            ws.cell(row=row, column=4, value="TOTAL:")
            ws.cell(row=row, column=4).font = Font(bold=True)
            ws.cell(row=row, column=5, value=total_emissions)
            ws.cell(row=row, column=5).font = Font(bold=True)

            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20

            wb.save(output_path)
            logger.info(f"Excel report generated: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return False

    @staticmethod
    def generate_pdf_report(project, calculations: List[Dict], output_path: Path) -> bool:
        """
        Generate PDF report for a project

        Args:
            project: Project object
            calculations: List of calculation dictionaries
            output_path: Output file path

        Returns:
            bool: True if successful
        """
        try:
            doc = SimpleDocTemplate(str(output_path), pagesize=A4)
            story = []
            styles = getSampleStyleSheet()

            # Title
            title = Paragraph(
                "GHG EMISSIONS REPORT",
                styles['Title']
            )
            story.append(title)
            story.append(Spacer(1, 0.3*inch))

            # Project info
            info_data = [
                ["Project ID:", str(project.id)],
                ["Project Name:", project.project_name],
                ["Organization:", project.organization_name],
                ["Reporting Year:", str(project.reporting_year)],
                ["Report Date:", datetime.now().strftime("%Y-%m-%d")],
                ["Status:", project.status]
            ]

            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
                ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))

            story.append(info_table)
            story.append(Spacer(1, 0.5*inch))

            # Emissions summary
            summary_title = Paragraph("EMISSIONS SUMMARY", styles['Heading2'])
            story.append(summary_title)
            story.append(Spacer(1, 0.2*inch))

            # Emissions table
            table_data = [["Scope", "Category", "Activity Data", "EF", "Emissions (tCO2e)"]]

            total_emissions = 0
            for calc in calculations:
                table_data.append([
                    calc.get('scope', 'N/A'),
                    calc.get('category', 'N/A'),
                    f"{calc.get('activity_data', 0):.2f}",
                    f"{calc.get('emission_factor', 0):.4f}",
                    f"{calc.get('emissions_tco2e', 0):.4f}"
                ])
                total_emissions += calc.get('emissions_tco2e', 0)

            # Add total row
            table_data.append(['', '', '', 'TOTAL:', f"{total_emissions:.4f}"])

            emissions_table = Table(table_data, colWidths=[1*inch, 2*inch, 1.2*inch, 1*inch, 1.5*inch])
            emissions_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
                ('FONT', (0, 1), (-1, -2), 'Helvetica', 9),
                ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(emissions_table)

            doc.build(story)
            logger.info(f"PDF report generated: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            return False

# Global instance
report_generator = ReportGenerator()
